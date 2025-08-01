import openpyxl
import pandas as pd
import os

#创建单点突变序列，同时储存到xlsx文件中
def build_matrix(input_str, fixed_str):
    x_len = len(input_str)
    y_len = len(fixed_str)

    matrix = []
    for row in range(x_len):
        row_list = []
        for col in range(y_len):
            new_str = input_str[:row] + fixed_str[col] + input_str[row + 1:]
            row_list.append(new_str)
        matrix.append(row_list)
    return matrix


def save_to_excel(matrix, filename="Single-point mutant peptide sequence.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DataMatrix"

    # 第一行写固定字母（Y轴）
    ws.cell(row=1, column=1, value="")  # 左上角空白
    for col_idx in range(len(matrix[0])):
        ws.cell(row=1, column=col_idx + 2, value=f"Y={col_idx + 1}")

    # 第一列写输入字母索引（X轴）
    for row_idx in range(len(matrix)):
        ws.cell(row=row_idx + 2, column=1, value=f"X={row_idx + 1}")

    # 写数据矩阵，左上角对应(1,1)
    for i, row in enumerate(matrix):
        for j, val in enumerate(row):
            ws.cell(row=i + 2, column=j + 2, value=val)

    wb.save(filename)
    print(f"数据已保存到 {filename}")


if __name__ == "__main__":
    input_str = input("输入多肽序列（如GAVL）：").strip().upper()
    # 定义合法选项
    valid_inputs = {'0', '1', '2', '3', '4'}
    while True:
        print("请选择二级结构类型（输入数字）:")
        print("0：默认")
        print("1：α 螺旋")
        print("2：反平行 β")
        print("3：平行 β")
        print("4：平坦")

        ss_input = input("请输入对应的数字（0-4）: ").strip()

        if ss_input in valid_inputs:
            ss_type = int(ss_input)
            break
        else:
            print("❌ 输入无效，请重新输入 0-4\n")
    fixed_str = "APGVMILFYWHCSTNQDEKR"
    matrix = build_matrix(input_str, fixed_str)
    save_to_excel(matrix)


# 读取xlsx种的序列
current_dir = os.path.dirname(os.path.abspath(__file__))
excel_file = os.path.join(current_dir, "Single-point mutant peptide sequence.xlsx")
df = pd.read_excel(excel_file, header=0)
# 跳过首行（列名）和首列（索引列）
data_matrix = df.iloc[1:, 1:].to_numpy()

# 打印矩阵查看,检查序列是否正确
# print(data_matrix)
# value = data_matrix[1, 2]  # 注意索引从0开始
# print("矩阵第2行第3列的值是：", value)

# 写入 PML 文件
pml_file = os.path.join(current_dir, "Somps.pml")
with open(pml_file, 'w') as f:
    for i in range(data_matrix.shape[0]):  # 行
        for j in range(data_matrix.shape[1]):  # 列
            sequence = data_matrix[i, j]
            if pd.notna(sequence):  # 跳过空值
                name = f"{i + 2}-{j + 2}"  # +2 是因为原始矩阵跳过首行首列
                f.write(f"fab {sequence}, name={name}, ss={ss_type}\n")

    # 写入最后一行保存命令
    f.write("save C:/Spmps.sdf\n")

print("PML 文件已生成：Somps.pml")


